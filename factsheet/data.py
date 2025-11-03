from datetime import date
from typing import List
from pathlib import Path
import os
import json
import pandas as pd
import yfinance as yf
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def canonical_symbol(symbol: str | None) -> str | None:
    if symbol is None:
        return None
    sym = str(symbol).strip()
    if not sym or sym.lower() == "nan":
        return None
    sym = sym.split(":")[0]
    sym = sym.replace(" ", "")
    if not sym:
        return None
    return sym.upper()


def prices_from_yfinance(tickers: List[str], start: date, end: date) -> pd.DataFrame:
    data = yf.download(tickers, start=start, end=end, auto_adjust=True, progress=False)
    if isinstance(data, pd.DataFrame) and 'Adj Close' in data.columns:
        adj = data['Adj Close']
    else:
        adj = data
    if isinstance(adj.columns, pd.MultiIndex):
        adj.columns = adj.columns.get_level_values(0)
    return adj.sort_index()


def fx_from_yfinance(base: str, quote: str, start: date, end: date) -> pd.Series:
    """Fetch daily FX rates via yfinance for base/quote pair (e.g., USD/ZMW)."""
    pair = f"{base}{quote}=X".upper()
    data = yf.download(pair, start=start, end=end + pd.Timedelta(days=1), progress=False, auto_adjust=False)
    if data.empty:
        return pd.Series(dtype=float)
    if isinstance(data.columns, pd.MultiIndex):
        close = data['Close']
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
    else:
        close = data['Close'] if 'Close' in data.columns else data.iloc[:, 0]
    close = close.sort_index()
    close.name = quote
    return close


def load_benchmark_prices(ticker: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.Series:
    """Load daily benchmark closing prices, ensuring coverage from start through end."""
    if pd.isna(start) or pd.isna(end):
        raise ValueError("Invalid start/end for benchmark download")
    buffer_start = start - pd.Timedelta(days=7)
    buffer_end = end + pd.Timedelta(days=1)
    data = yf.download(ticker, start=buffer_start, end=buffer_end, progress=False, auto_adjust=False)
    if data.empty:
        raise ValueError(f"No market data returned for {ticker}")

    if isinstance(data.columns, pd.MultiIndex):
        if 'Adj Close' in data.columns.get_level_values(0):
            close = data['Adj Close']
        else:
            close = data['Close']
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
    else:
        if 'Adj Close' in data.columns:
            close = data['Adj Close']
        else:
            close = data['Close'] if 'Close' in data.columns else data.iloc[:, 0]

    close = close.sort_index()
    close = close.loc[:end]
    full_index = pd.date_range(start=buffer_start, end=end, freq="D")
    close = close.reindex(full_index).ffill()

    valid_idx = close.index[close.index <= start]
    if len(valid_idx) == 0:
        raise ValueError(f"Benchmark {ticker} has no price on or before start date")

    aligned = close.loc[start:end]
    aligned.name = ticker
    return aligned


def load_account_value_timeseries(path: str) -> pd.Series:
    """Load account value time series (base currency) from Performance sheet."""
    df = pd.read_excel(path, sheet_name='Performance')

    date_col = None
    for cand in ['Date', 'date']:
        if cand in df.columns:
            date_col = cand
            break
    if date_col is None:
        raise ValueError("No 'Date' column found in Performance sheet")

    value_col = None
    candidates = [
        'AccountValueTimeSeries',
        'Account Value Time Series',
        'Account_Value_TimeSeries',
    ]
    for cand in candidates:
        if cand in df.columns:
            value_col = cand
            break
    if value_col is None:
        raise ValueError("No account value time series column found in Performance sheet")

    sub = df[[date_col, value_col]].dropna()
    sub[date_col] = pd.to_datetime(sub[date_col], dayfirst=True, errors='coerce')
    sub = sub.dropna(subset=[date_col])
    vals = pd.to_numeric(sub[value_col], errors='coerce')
    out = pd.Series(vals.values, index=sub[date_col].values).sort_index()
    out = out.dropna()
    out = out[~out.index.duplicated(keep='first')]
    return out


def load_benchmark_cumulative_returns(ticker: str, start: pd.Timestamp, end: pd.Timestamp) -> pd.Series:
    """Load benchmark cumulative returns (in decimal) between start and end dates using closing prices."""
    if pd.isna(start) or pd.isna(end):
        raise ValueError("Invalid start/end for benchmark download")
    buffer_start = start - pd.Timedelta(days=7)
    buffer_end = end + pd.Timedelta(days=1)
    data = yf.download(ticker, start=buffer_start, end=buffer_end, progress=False, auto_adjust=False)
    if data.empty:
        raise ValueError(f"No market data returned for {ticker}")

    if isinstance(data.columns, pd.MultiIndex):
        close = data['Close']
        if isinstance(close, pd.DataFrame):
            close = close.iloc[:, 0]
    else:
        close = data['Close'] if 'Close' in data.columns else data.iloc[:, 0]

    close = close.sort_index()
    close = close.loc[:end]
    if close.empty:
        raise ValueError(f"Benchmark {ticker} has no data available up to end date")

    full_index = pd.date_range(start=buffer_start, end=end, freq="D")
    close = close.reindex(full_index).ffill()

    valid_idx = close.index[close.index <= start]
    if len(valid_idx) == 0:
        raise ValueError(f"Benchmark {ticker} has no price on or before start date")
    baseline_price = close.loc[valid_idx[-1]]
    if pd.isna(baseline_price) or baseline_price == 0:
        raise ValueError(f"Benchmark {ticker} baseline price missing or zero")

    aligned = close.loc[start:end]
    returns = aligned / baseline_price - 1.0
    returns.iloc[0] = 0.0
    returns.name = ticker
    return returns


def load_positions_from_excel(path: str, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name)


def load_transactions_from_excel(path: str, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name)


def write_table_to_excel(path: str, sheet_name: str, top_left_cell: str, df: pd.DataFrame, named_range: str | None = None) -> None:
    """Write a DataFrame to an Excel sheet at the given top-left cell.

    If the workbook or sheet does not exist, they will be created.
    Optionally define a named range that spans the written table (including header).
    """
    try:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        wb = load_workbook(path)
    except FileNotFoundError:
        wb = Workbook()
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Clear existing contents (simple approach: delete and recreate sheet)
        idx = wb.sheetnames.index(sheet_name)
        wb.remove(ws)
        ws = wb.create_sheet(title=sheet_name, index=idx)
    else:
        ws = wb.create_sheet(title=sheet_name)

    # Write header
    start_col = ord(top_left_cell[0].upper()) - 64
    start_row = int(''.join(filter(str.isdigit, top_left_cell)))
    for j, col in enumerate(df.columns, start=start_col):
        ws.cell(row=start_row, column=j, value=str(col))
    # Write rows
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=val)

    # Named range covering header + data
    if named_range:
        end_row = start_row + len(df)
        end_col = start_col + len(df.columns) - 1
        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)
        ref = f"{sheet_name}!${start_col_letter}${start_row}:${end_col_letter}${end_row}"
        # Remove existing defined name if present
        for dn in list(wb.defined_names):
            if dn.name == named_range:
                wb.defined_names.delete(dn.name)
        wb.create_named_range(named_range, ws, ref.split('!')[1])

    # Ensure first sheet is not the default empty sheet for new workbooks
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) > 1:
        del wb['Sheet']

    wb.save(path)


def load_latest_holdings_from_aggregated(path: str, top_n: int | None = None) -> pd.DataFrame:
    """Return holdings (name, symbol, weight) from latest date in Aggregated Amounts sheet."""
    df = pd.read_excel(
        path,
        sheet_name="Aggregated Amounts",
        usecols=[
            "Date",
            "Instrument Description",
            "Instrument Symbol",
            "Underlying Instrument Symbol",
            "Amount Client Currency",
            "Amount Type Name",
        ],
    )
    df["Date"] = pd.to_datetime(df["Date"], dayfirst=True, errors="coerce")
    df = df.dropna(subset=["Date", "Amount Client Currency"])
    if df.empty:
        raise ValueError("Aggregated Amounts sheet missing required data")

    latest_date = df["Date"].max()
    latest = df[df["Date"] == latest_date].copy()
    if latest.empty:
        raise ValueError("No rows for latest date in Aggregated Amounts")

    latest["Amount Type Name"] = latest["Amount Type Name"].astype(str)
    cash_mask = latest["Amount Type Name"].str.contains("cash", case=False, na=False)

    non_cash = latest[~cash_mask & latest["Instrument Description"].notna()].copy()
    cash_rows = latest[cash_mask].copy()
    if not cash_rows.empty:
        cash_rows["Instrument Description"] = "Cash"
        cash_rows["Instrument Symbol"] = "CASH"
        cash_rows["Underlying Instrument Symbol"] = "CASH"
    latest = pd.concat([non_cash, cash_rows], ignore_index=True)

    latest["symbol"] = (
        latest["Instrument Symbol"]
        .fillna(latest["Underlying Instrument Symbol"])
        .fillna(latest["Instrument Description"])
        .astype(str)
        .str.strip()
    )
    latest["symbol"] = latest["symbol"].str.split(":").str[0].str.strip()
    latest["symbol"] = latest["symbol"].apply(canonical_symbol)
    latest["name_display"] = (
        latest["Instrument Description"]
        .fillna(latest["Instrument Symbol"])
        .fillna(latest["Underlying Instrument Symbol"])
        .fillna(latest["symbol"])
        .astype(str)
        .str.strip()
    )
    latest.loc[latest["symbol"].isna() | (latest["symbol"] == ""), "symbol"] = "CASH"
    latest.loc[
        latest["name_display"].isna() | (latest["name_display"].str.lower() == "nan") | (latest["name_display"].str.strip() == ""),
        "name_display",
    ] = "Cash"

    grouped = (
        latest.groupby("symbol", as_index=False)
        .agg(
            Amount=("Amount Client Currency", "sum"),
            name=("name_display", "first"),
        )
        .sort_values("Amount", ascending=False)
    )
    grouped = grouped[grouped["Amount"] > 0]
    if grouped.empty:
        raise ValueError("No positive holdings found on latest date")

    total = grouped["Amount"].sum()
    grouped["weight"] = grouped["Amount"] / total
    grouped["Date"] = latest_date
    result = grouped.rename(columns={"symbol": "symbol"})
    result = result[["Date", "name", "symbol", "weight"]]
    if top_n is not None:
        result = result.head(top_n)
    return result.reset_index(drop=True)


def load_symbol_overrides(path: str | None) -> dict:
    overrides: dict[str, dict] = {}
    if not path:
        return overrides
    file_path = Path(path)
    if not file_path.is_file():
        return overrides
    df = pd.read_csv(file_path)
    for _, row in df.iterrows():
        entry: dict[str, str] = {}
        sector = row.get("sector")
        region = row.get("region")
        country = row.get("country")
        yf_symbol = row.get("yf_symbol")
        if pd.notna(sector) and str(sector).strip():
            entry["sector"] = str(sector).strip()
        if pd.notna(region) and str(region).strip():
            entry["region"] = str(region).strip()
        if pd.notna(country) and str(country).strip():
            entry["country"] = str(country).strip()
        if pd.notna(yf_symbol) and str(yf_symbol).strip():
            entry["yf_symbol"] = canonical_symbol(yf_symbol)
        if not entry:
            continue
        sym_key = canonical_symbol(row.get("symbol"))
        name_key = str(row.get("name") or "").strip().upper()
        keys = []
        if sym_key:
            keys.append(sym_key)
        if name_key:
            keys.append(name_key)
        for key in keys:
            overrides[key] = {**entry}
    return overrides


def load_country_region_map(path: str | None) -> dict:
    mapping = {}
    if not path:
        return mapping
    file_path = Path(path)
    if not file_path.is_file():
        return mapping
    df = pd.read_csv(file_path)
    for _, row in df.iterrows():
        country = str(row.get("country") or "").strip()
        region = str(row.get("region") or "").strip()
        if country and region:
            mapping[country.lower()] = region
    return mapping


def load_sector_lookup(path: str | None) -> dict:
    lookup: dict[str, dict] = {}
    if not path:
        return lookup
    file_path = Path(path)
    if not file_path.is_file():
        return lookup
    df = pd.read_csv(file_path)
    for _, row in df.iterrows():
        source_type = str(row.get("source_type") or "").strip().lower()
        source_value = str(row.get("source_value") or "").strip().lower()
        canonical = str(row.get("canonical_sector") or "").strip()
        if not source_type or not source_value or not canonical:
            continue
        lookup.setdefault(source_type, {})[source_value] = canonical
    return lookup


def load_classification_cache(path: str | None) -> dict:
    if not path:
        return {}
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
    if not isinstance(raw, dict):
        return {}
    cache: dict[str, dict] = {}
    for key, val in raw.items():
        canon = canonical_symbol(key) or str(key).strip().upper()
        if not canon:
            continue
        entry = {}
        if isinstance(val, dict):
            for field in ("sector", "region", "country", "yf_symbol", "industry", "error"):
                if field in val and val[field] is not None:
                    entry[field] = val[field]
        cache[canon] = entry
    return cache


def save_classification_cache(cache: dict, path: str | None) -> None:
    if not path:
        return
    try:
        serializable = {}
        for key, val in cache.items():
            canon = canonical_symbol(key) or str(key).strip().upper()
            if not canon or not isinstance(val, dict):
                continue
            serializable[canon] = val
        with open(path, "w", encoding="utf-8") as f:
            json.dump(serializable, f, indent=2, sort_keys=True)
    except OSError:
        pass


def resolve_symbol_classification(symbol: str | None,
                                  name: str | None,
                                  overrides: dict,
                                  cache: dict,
                                  sector_lookup: dict) -> tuple[str | None, str | None, str | None, str | None, str | None, str | None, str | None, bool]:
    cache_changed = False
    symbol_key = canonical_symbol(symbol)
    name_str = str(name).strip() if name is not None else ""
    name_key = name_str.upper() if name_str else ""

    if symbol_key in {"CASH", "USD", "CASHUSD"} or (name_key and "CASH" in name_key):
        return "Cash", None, None, "Cash", "Cash", "Cash", None, False

    search_keys = []
    if symbol_key:
        search_keys.append(symbol_key)
    if name_key:
        search_keys.append(name_key)

    sector_raw = None
    industry_raw = None
    country_raw = None
    yf_symbol_override = None
    override = None
    for key in search_keys:
        if key in overrides:
            override = overrides[key]
            break

    override_sector = override.get("sector") if override else None
    override_region = override.get("region") if override else None
    override_country = override.get("country") if override else None
    if override:
        yf_symbol_override = override.get("yf_symbol")

    cache_key = symbol_key or name_key or None
    cache_entry = cache.get(cache_key, {}) if cache_key else {}
    if cache_entry:
        sector_raw = cache_entry.get("sector")
        industry_raw = cache_entry.get("industry")
        country_raw = cache_entry.get("country")
        yf_symbol_override = yf_symbol_override or cache_entry.get("yf_symbol")
        cached_error = bool(cache_entry.get("error"))
    else:
        cached_error = False

    fetch_symbol = yf_symbol_override or symbol_key
    fetch_needed = bool(fetch_symbol) and (sector_raw is None or country_raw is None or industry_raw is None) and not cached_error
    if fetch_needed:
        try:
            info = yf.Ticker(fetch_symbol).info
        except Exception:
            info = {}
        fetched_sector = info.get("sector") or info.get("industry")
        fetched_industry = info.get("industry")
        fetched_country = info.get("country")
        if cache_key:
            cache_entry = cache.setdefault(cache_key, {})
            if fetched_sector is not None:
                cache_entry["sector"] = fetched_sector
            if fetched_industry is not None:
                cache_entry["industry"] = fetched_industry
            if fetched_country is not None:
                cache_entry["country"] = fetched_country
            if fetch_symbol is not None:
                cache_entry["yf_symbol"] = fetch_symbol
            cache_entry.pop("error", None)
            cache_changed = True
        if sector_raw is None:
            sector_raw = fetched_sector
        if industry_raw is None:
            industry_raw = fetched_industry
        if country_raw is None:
            country_raw = fetched_country

    if (sector_raw is None and country_raw is None and industry_raw is None) and fetch_symbol and cache_key:
        cache_entry = cache.setdefault(cache_key, {})
        if not cache_entry.get("error"):
            cache_entry["error"] = True
            cache_changed = True

    return (
        sector_raw,
        industry_raw,
        country_raw,
        override_sector,
        override_region,
        override_country,
        yf_symbol_override,
        cache_changed,
    )


def load_daily_portfolio_returns_from_performance(path: str) -> pd.Series:
    """Load daily % returns from 'Performance' sheet.

    Expects a column named like '% daily returns' and a date column named 'Date'.
    Returns a pandas Series indexed by datetime with decimal returns (e.g., 0.0123 for 1.23%).
    """
    df = pd.read_excel(path, sheet_name='Performance')
    # Try common column name variants
    date_col = None
    for cand in ['Date', 'date']:
        if cand in df.columns:
            date_col = cand
            break
    if date_col is None:
        raise ValueError("No 'Date' column found in Performance sheet")

    ret_col = None
    for cand in ['% daily returns', 'daily returns %', 'Daily % return', 'Daily return %']:
        if cand in df.columns:
            ret_col = cand
            break
    if ret_col is None:
        # Heuristic: first column containing '%' and 'return'
        for c in df.columns:
            lc = str(c).lower()
            if '%' in lc and 'return' in lc:
                ret_col = c
                break
    if ret_col is None:
        raise ValueError("No daily return percentage column found in Performance sheet")

    s = df[[date_col, ret_col]].dropna()
    s[date_col] = pd.to_datetime(s[date_col], dayfirst=True, errors='coerce')
    s = s.dropna(subset=[date_col])
    # Convert from strings like '1.23 %' or numbers in percent to decimals
    vals_raw = pd.to_numeric(
        s[ret_col].astype(str).str.replace('%', '', regex=False).str.replace(' ', '', regex=False),
        errors='coerce',
    )
    if vals_raw.abs().max() > 1:
        vals = vals_raw / 100.0
    else:
        vals = vals_raw
    out = pd.Series(vals.values, index=s[date_col].values).sort_index()
    out = out[~out.index.duplicated(keep='first')]
    return out


def load_accumulated_twr_from_performance(path: str) -> pd.Series:
    """Load cumulative (% absolute) TWR series from 'Performance' sheet.

    Returns a pandas Series indexed by datetime with decimal returns
    (e.g., 1 -> 0.01).
    """
    df = pd.read_excel(path, sheet_name='Performance')

    date_col = None
    for cand in ['Date', 'date']:
        if cand in df.columns:
            date_col = cand
            break
    if date_col is None:
        raise ValueError("No 'Date' column found in Performance sheet")

    value_col = None
    candidates = [
        'AccumulatedTimeWeightedTimeSeries',
        'Accumulated Time Weighted Time Series',
        'Accumulated_TWR',
        'Accumulated TWR',
    ]
    for cand in candidates:
        if cand in df.columns:
            value_col = cand
            break
    if value_col is None:
        raise ValueError("No accumulated TWR column found in Performance sheet")

    sub = df[[date_col, value_col]].dropna()
    sub[date_col] = pd.to_datetime(sub[date_col], dayfirst=True, errors='coerce')
    sub = sub.dropna(subset=[date_col])
    vals = pd.to_numeric(sub[value_col], errors='coerce')
    vals = vals / 100.0
    out = pd.Series(vals.values, index=sub[date_col].values).sort_index()
    out = out[~out.index.duplicated(keep='first')]
    return out



